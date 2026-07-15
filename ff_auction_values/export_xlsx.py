#!/usr/bin/env python3
"""
Convert auction_values_half_ppr.csv into a formatted .xlsx for easy viewing
(frozen header row, sane column widths, position-colored rows) — same data,
no calculations, just presentation.

Usage:
    python export_xlsx.py
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
CSV_FILE = HERE / "auction_values_half_ppr.csv"
XLSX_FILE = HERE / "auction_values_half_ppr.xlsx"

POSITION_FILL = {
    "QB": "FFE0B2",
    "RB": "C8E6C9",
    "WR": "BBDEFB",
    "TE": "F8BBD0",
}


def main():
    with open(CSV_FILE, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    header, data = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Auction Values (Half-PPR)"

    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    pos_col = header.index("Position")
    numeric_cols = {header.index(c) for c in
                    ("Overall Rank", "Position Rank", "Proj Rank (Pos)",
                     "Personal Rank (Pos)", "Blended Points", "VORP",
                     "Auction Value ($)")}

    for row in data:
        out_row = []
        for i, val in enumerate(row):
            if i in numeric_cols and val != "":
                out_row.append(float(val) if "." in val else int(val))
            else:
                out_row.append(val)
        ws.append(out_row)
        fill_color = POSITION_FILL.get(row[pos_col])
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill

    widths = {"Overall Rank": 8, "Player": 24, "Position": 9, "Team": 7,
              "Position Rank": 9, "Proj Rank (Pos)": 10, "Personal Rank (Pos)": 12,
              "Blended Points": 12, "VORP": 9, "Auction Value ($)": 12}
    for i, col_name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 12)

    ws.auto_filter.ref = ws.dimensions

    wb.save(XLSX_FILE)
    print(f"Wrote {len(data)} players to {XLSX_FILE}")


if __name__ == "__main__":
    main()
