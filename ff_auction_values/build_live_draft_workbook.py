#!/usr/bin/env python3
"""
Build a standalone Excel workbook that recalculates auction values live as
a real draft happens. New file, no legacy formatting to preserve, so this
uses openpyxl (not win32com like the Cheat Sheet workbook) — simpler, and
doesn't require Excel to be running.

Sheets:
  Setup        - league settings. Budget, Number of Teams, Scoring Format,
                 AND now Roster Shape (QB/RB/WR/TE starters; FOUR
                 independent flex-type counts — RB/WR/TE, RB/WR, WR/TE,
                 Superflex — that combine freely, e.g. 1 RB/WR + 1 WR/TE +
                 1 Superflex all in the same league; DEF, K, Bench) are all
                 live-editable. Teams and Format are restricted dropdowns
                 (real market-data calibration exists for exactly those
                 combinations — see CLAUDE.md); Roster Shape inputs are
                 also bounded dropdowns (QB/RB/WR/TE starters 0 up to the
                 project's tested max; each flex-type count 0-6) but work
                 by LIVE FORMULA adjustment on top of the nearest real
                 Teams/Format baseline, not a precomputed lookup — see
                 "v2: roster shape" below for why that's a different
                 mechanism than Teams/Format use.
  Draft Board  - one column-block per position (QB / RB / WR / TE side by
                 side). Enter the real price paid as players are drafted;
                 Live Value updates automatically for everyone still on
                 the board.
  Live Calc    - the actual recalculation math.
  My Team      - personal budget tracker, unrelated to the live recalc
                 math (see add_my_team_sheet()).

## v2: roster shape — a live formula, not a precomputed lookup, and why

Team count and scoring format are each a small, fully enumerable set (5
team counts x 3 formats = 15) with real market data for every single
combination — so v1.4 precomputed Weighted VORP for all 15 and had Excel
pick among them. Roster shape is different: the bounded starter ranges
alone, crossed with four independently-combinable flex-type counts, is
far more combinations than could ever get real data, and pulling real
data for each isn't feasible (see CLAUDE.md, "Round 7" for the full
account of why real data is needed at all rather than a pure formula —
three separate proven cases, not a stylistic choice).

So instead: REPLACEMENT_RANK becomes a LIVE FORMULA
(`roster_formula.py`'s model, replicated in Excel), anchored to whichever
(Teams, Format) baseline is already real-calibrated:
    demand(pos) = starters[pos] + SUM over the 4 independent flex-type
                  counts of (that type's flex_share[pos] * that type's count)
    adjusted_rank(pos) = baseline_rank(pos) * demand(pos) / demand(pos; BASELINE roster)
    replacement_points(pos) = LARGE(that position's full points column, adjusted_rank(pos))
    weighted_vorp(player) = MAX(0, player_points - replacement_points(pos)) ^ exponent(pos)

`LARGE(range, k)` is a single native Excel function returning "the k-th
largest value" — this is what makes a live version of Harstad-style
replacement-level lookup safe to build: no custom interpolation formula,
no array/dynamic-array construct (the category that caused a real file-
corruption problem earlier in this project — see the "Big Board" postmortem
below). Every player's raw points (one column per scoring format) sit
directly in the sheet for this to reference.

VORP_EXPONENT and POSITION_BUDGET_SHARE do NOT use this smooth formula —
tested directly and found to fail by 2.5x for the one case that matters
most (superflex QB budget share, see CLAUDE.md). Both use a REGIME SWITCH
instead: held at the real (Teams, Format) baseline when there's no
2-QB-equivalent demand (QB starters = 1 and superflex count = 0), or
shifted by the REAL observed superflex ratio (measured once, at 12-team/
half-PPR — calibrate_superflex.py) when QB starters = 2 OR superflex
count >= 1. This is an extrapolation for any (Teams, Format) other than
12-team/half-PPR itself, clearly weaker evidence than the smooth rank
formula, but still anchored to real data rather than assumed.

**Tier is simplified in this version.** The natural-gap-detection
algorithm (tiers.py) can't be replicated as a live Excel formula without
the same array-formula risk noted above. Replaced with fixed-width RANK
bands (every 5 players by Weighted VORP = one tier) — a real, disclosed
downgrade from the adaptive gap-detection tiering, not a bug. Still gives
localized (not just whole-draft) live re-rating, just with mechanically
fixed band sizes instead of detected natural breaks.

## Live recalibration (unchanged since v1.2): a two-factor model

1. TIER-LEVEL market re-rating (same direction, local):
      tier_factor = SUM(real prices paid so far in this tier)
                    / SUM(those same players' original static values)
      held at 1.0 until >= TIER_MIN_SAMPLE players in the tier are drafted,
      then clamped to [TIER_FACTOR_MIN, TIER_FACTOR_MAX]
2. WHOLE-DRAFT-LEVEL budget depletion (global, all positions combined):
      global_factor = current_whole_draft_rate / initial_whole_draft_rate
      clamped to [GLOBAL_FACTOR_MIN, GLOBAL_FACTOR_MAX]
Final: live_value = ROUND(static_value * tier_factor * global_factor, 0)
       static_value = ROUND(1 + weighted_vorp * position_rate, 0)

**Bug: unbounded tier re-rating (fixed).** Both factors above were originally
uncapped ratios. Confirmed via direct testing: because TIER_BAND_SIZE=5 fixed
bands don't account for value composition, roughly two-thirds of a typical
160+-deep RB board sits at the $1 static-value floor, and any tier straddling
that floor mixes near-zero-differentiation players with slightly-better ones
(e.g. a real tier: statics $4, $3, $2, $2, $1). A single $1->$5 winning bid on
the $1 player produced tier_factor = 5/1 = 5.0, applied uniformly to the whole
tier -- the $4 player's live value jumped to $20. This is what the user
reported as "the rest of the $1 players in his tier all changed to $5" (the
actual effect was worse: non-floor tier-mates moved too). MEDIAN(lo, hi, x) now
clamps both factors so one anomalous bid can locally re-rate a tier/the whole
board within a bounded range instead of by an arbitrary multiple. Bounds were
chosen so normal fair-value drafting (factor stays near 1.0) is completely
unaffected, and even a wild $150+ single-player overpay (previously observed
to move the global factor to ~0.93) still lands safely inside [0.7, 1.3].

**Follow-up: single-bid tier re-rating (fixed).** Even bounded, letting one
drafted player re-rate the rest of their tier was still wrong on its own
terms: one bid is one bidder's opinion on one specific player, not evidence
about the whole tier, and the original fixed-5-player bands didn't guarantee
tier-mates were actually similar in value (see the $4/$3/$2/$2/$1 example
above -- that's one tier). Fix: tier_factor now stays at exactly 1.0 (no
re-rating) until TIER_MIN_SAMPLE=2 players in that tier have been drafted --
a second data point is the minimum needed to distinguish "a trend in this
tier" from "one idiosyncratic price."

**Round 8 extension: value-proportional tiers, superseded.** The
fixed-5-player-band Tier assignment was first replaced with a cumulative-
value-share bucket scheme (a new tier every time running Weighted VORP
crossed another 5% slice of the position's total) to fix a real complaint:
Bijan Robinson and Christian McCaffrey landed in the same tier despite an
~90-point Weighted VORP gap between them. This shipped in two bad versions
in a row -- an inclusive-share/CEILING version where the single most
valuable player's own share regularly exceeded one whole band by itself
(Gibbs alone was ~6.2% of the RB pool), making tier 1 structurally
unreachable at every position; and, after excluding each player's own
share fixed that, a related edge case where every at-or-below-replacement
player (Weighted VORP = 0) computed a share of exactly 1.0 and overflowed
one tier past the table's sized range, producing 363 real formula errors.
Both are now moot -- see the gap-detection design below, which replaced
the whole bucket-based approach.

**Round 8 fourth extension: gap-detection tiers.** The bucket-based design
above was fundamentally the wrong shape for the problem: a boundary drawn
every fixed X% of cumulative value doesn't track where the real gaps in
the curve actually are, so two nearly-identical players landing on
opposite sides of a boundary (Gibbs/Bijan, 0.8% apart in Weighted VORP)
get needlessly split, while the exact same mechanism can (for QB, with a
smaller/steeper pool) let one huge single-player share jump the running
total past more than one bucket in a single step, silently skipping a
tier number. User caught both independently, correctly, in the same
message: "personally I would put Gibbs and Bijan in the same tier...
Why are there missing tiers in QB, I would think it would also go in
incremental order no matter what the separation is." Replaced entirely
with actual nearest-neighbor gap detection:
    Tier Start(i) = 1 if i is the canonical (lowest-Row-Num) player at its
                    own Weighted VORP AND EITHER no one ranks above it OR
                    the relative drop from the player immediately above
                    exceeds GAP_THRESHOLD; else 0
    Tier(i) = COUNTIFS(wvorp_range, ">="&own_wvorp, tier_start_range, 1)
Tier is a running COUNT of how many real gap-crossings occur at or above
this player, so tier numbers are consecutive by construction -- no skips
are possible regardless of how large any individual gap is, which
structurally fixes the QB complaint as a side effect of fixing the first
one. GAP_THRESHOLD=8% was chosen by testing candidate thresholds (5/6/8/
10/12%) against real Weighted VORP data until one cleanly separated every
case the user named: Gibbs/Bijan (0.8% gap) and Puka Nacua/Ja'Marr Chase
and Trey McBride/Brock Bowers (all near-zero gaps) stay together; Bijan/
McCaffrey (14.6%) and Jonathan Taylor/James Cook (21.9%) split apart.
"Row Num" (a literal integer written once at Python build time, not a
live =ROW() formula) exists specifically to break ties correctly: many
players legitimately share the exact same Weighted VORP (most commonly
0, at or below replacement -- confirmed 100+ players tied at zero on a
160-deep RB board), and MINIFS(">"&own) skips over tied values to the
next DISTINCT one, so without a canonical-representative check every
tied player would independently see the same gap and each flag as its
own tier start, fragmenting what should be one shared tier into dozens.
Tier count is no longer a fixed, Setup-independent constant (it depends
on how many real gaps the data has, which can vary by roster config) --
the Live Calc tier table is sized to each position's own n_rows, the
true worst case, deliberately generous rather than clever given this
project has already hit the "#N/A because the lookup table was too
small" failure twice in this same round.

**Round 8 extension: tier-factor value dampening.** Even with the
TIER_MIN_SAMPLE gate, user pushback (correct): "even if two players are
overpaid for in a $1 tier, I'm not sure the [other] $3 player should almost
double... that value really shouldn't change much." A tier's factor is now
also scaled by TIER_DAMPEN_BUDGET_FRACTION (10% of Setup!Budget, e.g. $20 in
a $200 league): tiers averaging at/above that dollar level get the full
re-rating signal, tiers below it get proportionally less -- a tier averaging
$2.40 (the $4/$3/$2/$2/$1 example) gets ~12% of the raw signal, so two real
overpays there now move the group a couple dollars, not double it. Singleton
tiers (common at the top under gap detection, whenever a player has no
real neighbor within GAP_THRESHOLD) can never reach TIER_MIN_SAMPLE=2 at
all, so elite, one-of-a-kind players only ever move via the whole-draft
global factor, never the tier factor -- which is the correct outcome once
tiers are precise enough that a true stud has no real tier-mate to compare
against.

Usage:
    python build_live_draft_workbook.py
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from build_auction_values import (
    POINTS_COL, STARTERS, FLEX_SLOTS, BENCH_SPOTS, BUDGET, TEAMS,
    NON_SKILL_SLOTS_PER_TEAM,
    blend_with_personal_ranks, load_personal_ranks, load_projections,
)
from build_teamcount_estimate import CALIBRATED, TEAM_COUNTS, FORMATS
from roster_formula import (
    FLEX_TYPES, SUPERFLEX_ANCHOR, BASELINE_STARTERS,
    EXPONENT_SLOPE_PER_FLEX, BUDGET_SHARE_SLOPE_PER_FLEX,
    STARTER_RANK_DAMPING,
)

OUT_FILE = Path(r"C:\Users\jbond\Dropbox\F6P Admin\Fantasy Football Cheat Sheet"
                r"\2026-Auction-Draft-Tool-working-copy.xlsx")

POSITION_FILL = {
    "QB": "FFE0B2", "RB": "C8E6C9", "WR": "BBDEFB", "TE": "F8BBD0",
}

FORMAT_LABEL = {"std": "STD", "half_ppr": "Half-PPR", "ppr": "PPR"}
DEFAULT_FORMAT = "half_ppr"
CONFIG_KEYS = [(fmt, teams) for fmt in FORMATS for teams in TEAM_COUNTS]

# Board depth generous enough for the deepest realistic roster shape in
# the bounded range (QB2, RB3, WR4, TE2, up to 8 total flex/superflex
# slots) — wider than v1.4's union-of-15-configs approach since row
# selection can no longer be based on precomputed weighted VORP per
# config (there's no longer a fixed set of configs). Uses raw points rank
# directly instead: top-N by blended points, per position, generous
# enough to comfortably cover the deepest bounded roster shape.
BOARD_DEPTH = {"QB": 60, "RB": 160, "WR": 190, "TE": 85}

TOTAL_ROSTER_SPOTS_PER_TEAM = (
    sum(STARTERS.values()) + FLEX_SLOTS + BENCH_SPOTS + NON_SKILL_SLOTS_PER_TEAM
)

# Draft Board columns per position. Far fewer than v1.4 (11 vs 37) since
# Weighted VORP is now computed live from 3 raw-points columns instead of
# precomputed per (format, teams) combination.
BLOCK_COLS = ["Player", "Team", "Points (STD)", "Points (Half-PPR)", "Points (PPR)",
              "Proj Pts", "Tier", "Weighted VORP", "Row Num",
              "Gap Start", "Gap Tier Top", "Tier Start",
              "Starting Value", "Price Paid", "Market Value"]
(OFF_PLAYER, OFF_TEAM, OFF_PTS_STD, OFF_PTS_HALF, OFF_PTS_PPR, OFF_SEL_PTS,
 OFF_TIER, OFF_WVORP, OFF_ROWNUM,
 OFF_GAP_START, OFF_GAPTIER_TOP, OFF_TIER_START,
 OFF_STATIC, OFF_PRICE, OFF_LIVE) = range(len(BLOCK_COLS))
BLOCK_WIDTH = len(BLOCK_COLS)
SPACER_COLS = 1

# Tier = gap detection with a cumulative-width cap on top (see module
# docstring, "Round 8 sixth extension: cumulative tier-width cap"). A new
# tier starts wherever the relative drop in Weighted VORP from the player
# immediately above exceeds GAP_THRESHOLD (no single neighbor gap is big
# enough on its own to matter) -- verified against real data: 8% cleanly
# keeps near-equal players together (Gibbs/Bijan 0.8% gap, Puka/Chase,
# McBride/Bowers) while separating real cliffs (Bijan/CMC 14.6%, Jonathan
# Taylor/James Cook 21.9%). But a long chain of small individual gaps can
# still add up to a wide total range with no internal boundary (confirmed:
# a real 15-player RB tier spanned Cook $39 down to Skattebo $25, a 56%
# swing, with no single gap over 7.6%) -- MAX_TIER_WIDTH additionally
# forces a split once cumulative range from a tier's own top exceeds this,
# even with no single big neighbor gap. 25% was chosen by testing 20/25/
# 30/35% against real data until it split the wide Cook-Skattebo chain
# into two reasonably-sized groups without touching any of the close pairs
# above.
GAP_THRESHOLD = 0.08
MAX_TIER_WIDTH = 0.25

# Bounds on the two multiplicative re-rating factors -- see "Bug: unbounded
# tier re-rating" in the module docstring. Without a cap, a single overpay
# on a near-replacement ($1-floor) player can multiply an entire tier by
# an arbitrary ratio (confirmed: a $1->$5 bid multiplied every player in
# that 5-player tier by 5x, including a $4 player that jumped to $20).
TIER_FACTOR_MIN, TIER_FACTOR_MAX = 0.5, 2.0
GLOBAL_FACTOR_MIN, GLOBAL_FACTOR_MAX = 0.7, 1.3

# A single drafted player is one bidder's opinion on one player, not a market
# trend -- a tier's factor stays at 1.0 (no re-rating) until at least this
# many players in it have been drafted. See CLAUDE.md "Round 8".
TIER_MIN_SAMPLE = 2

# How much a tier's re-rating factor is allowed to move depends on how much
# real dollar value is actually in that tier -- a tier of $1-floor players
# has no real differentiation to react to, so its factor is dampened toward
# 1.0; a tier of $20+ (10% of a $200 budget) players gets the full signal.
# Scales with Setup!Budget so it isn't wrong for very small/large leagues.
TIER_DAMPEN_BUDGET_FRACTION = 0.10


def compute_all_points():
    """Run the projection+personal-rank blend once per format. Returns
    {fmt: {pos: [{"name":,"team":,"blended_points":}, ...]}} sorted desc."""
    by_fmt = {}
    for fmt in FORMATS:
        projections = load_projections(fmt)
        personal_ranks = load_personal_ranks(fmt)
        by_fmt[fmt] = blend_with_personal_ranks(projections, personal_ranks)
    return by_fmt


def board_player_list(by_fmt):
    """Row list per position: union of each format's own top-N by blended
    points (points, not weighted VORP, since weighted VORP is no longer
    precomputed) — generous BOARD_DEPTH covers the realistic bounded
    roster-shape range. Initial row order (before any user sort, via the
    Table sort buttons) is DEFAULT_FORMAT's own points descending, so it
    matches what the "Proj Pts" column actually shows on load (Setup!Format
    defaults to DEFAULT_FORMAT) -- a player who only makes the cut via a
    different format's depth falls back to their max-across-formats value
    for sort purposes, since they have no DEFAULT_FORMAT points of their
    own to sort by."""
    from name_match import normalize_name

    by_pos = {pos: {} for pos in POINTS_COL}
    for fmt in FORMATS:
        for pos in POINTS_COL:
            top = by_fmt[fmt][pos][:BOARD_DEPTH[pos]]
            for p in top:
                k = normalize_name(p["name"])
                by_pos[pos][k] = max(by_pos[pos].get(k, 0.0), p["blended_points"])

    default_pts = {pos: {normalize_name(p["name"]): p["blended_points"]
                          for p in by_fmt[DEFAULT_FORMAT][pos]}
                   for pos in POINTS_COL}

    return {pos: sorted(by_pos[pos].keys(),
                         key=lambda k: -default_pts[pos].get(k, by_pos[pos][k]))
            for pos in POINTS_COL}


def build_block_layout(board_keys):
    layout = {}
    col = 1
    for pos in POINTS_COL:
        n = len(board_keys[pos])
        layout[pos] = {"start_col": col, "n_rows": n, "last_row": n + 1}
        col += BLOCK_WIDTH + SPACER_COLS
    return layout


def block_col(layout, pos, offset):
    return get_column_letter(layout[pos]["start_col"] + offset)


def block_range(layout, pos, offset):
    last_row = layout[pos]["last_row"]
    col = block_col(layout, pos, offset)
    return f"'Draft Board'!${col}$2:${col}${last_row}"


def teamfmt_choose(cell_refs_by_config):
    """=CHOOSE(index, ref1, ..., ref15) — index computed from BOTH
    Setup!Teams and Setup!Format (same pattern as v1.4)."""
    teams_list = "{" + ",".join(str(t) for t in TEAM_COUNTS) + "}"
    fmt_list = "{" + ",".join(f'"{FORMAT_LABEL[f]}"' for f in FORMATS) + "}"
    index = (f"MATCH(Setup!$B$5,{teams_list},0)"
             f"+(MATCH(Setup!$B$7,{fmt_list},0)-1)*{len(TEAM_COUNTS)}")
    refs = [cell_refs_by_config[key] for key in CONFIG_KEYS]
    return f"CHOOSE({index},{','.join(refs)})"


def add_instructions_sheet(wb):
    """Customer-facing walkthrough, styled to match the Cheat Sheet
    workbook's own Instructions tab (navy title banner, "Step N:" callouts,
    "Note" warnings, a Field Legend) -- kept as its own sheet rather than
    folded into Setup so the actual league inputs aren't pushed down the
    page by a wall of explanatory text."""
    sheet = wb.create_sheet("Instructions", 0)

    title_fill = PatternFill(start_color="660000", end_color="660000", fill_type="solid")
    title_font = Font(bold=False, size=22, color="FFFFFF")
    step_font = Font(bold=True, size=11)
    note_font = Font(bold=True, size=11)
    body_font = Font(bold=False, size=11)
    legend_header_font = Font(bold=True, size=12)

    sheet.merge_cells("A2:I2")
    sheet["A2"] = "Live Auction Draft Board"
    sheet["A2"].font = title_font
    sheet["A2"].fill = title_fill
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sheet.row_dimensions[2].height = 28.5

    def step(row, label, *lines):
        sheet[f"A{row}"] = label
        sheet[f"A{row}"].font = step_font
        for i, line in enumerate(lines):
            sheet[f"B{row + i}"] = line
            sheet[f"B{row + i}"].font = body_font
            sheet[f"B{row + i}"].alignment = Alignment(wrap_text=True, vertical="top")
        return row + len(lines) + 1

    r = 4
    r = step(r, "Step 1:",
              "Go to the Setup tab and enter your league's Starting Budget, "
              "Number of Teams, and Scoring Format.")
    r = step(r, "Step 2:",
              "Still on Setup, enter your Roster Shape -- Starters at each "
              "position, any Flex slots (RB/WR/TE, RB/WR, WR/TE, or "
              "Superflex -- these can combine, e.g. 1 RB/WR flex AND 1 "
              "Superflex in the same league), DEF, K, and Bench spots.",
              "Every value on the whole board recalculates instantly when "
              "you change anything here -- there's nothing to re-run.")
    r = step(r, "Step 3:",
              "Go to the Draft Board tab. Each position (QB/RB/WR/TE) has "
              "its own set of columns. \"Starting Value\" is what a player "
              "is worth before any picks are made. \"Market Value\" is the "
              "live price that updates as your draft happens.")
    r = step(r, "Step 4:",
              "As players get drafted -- by you or anyone else in your "
              "league -- enter the actual winning bid in that player's "
              "\"Price Paid\" column. You don't need to delete, sort, or "
              "remove anything; Market Value for every remaining player "
              "updates on its own.",
              "Click the dropdown arrow on any position's header row to "
              "sort that block -- handy after switching Scoring Format, "
              "since row order is only set once when the sheet is built.")
    r = step(r, "Step 5:",
              "Track your own roster and remaining budget on the My Team "
              "tab -- separate from the main board, just for you.")
    r += 1

    sheet[f"A{r}"] = "Note"
    sheet[f"A{r}"].font = note_font
    sheet[f"B{r}"] = ("A single pick won't swing anyone's price much -- Market "
                       "Value only moves once a couple of real prices come in "
                       "for players near each other in value, and even then "
                       "it's a modest nudge, not a big swing.")
    sheet[f"B{r}"].font = body_font
    sheet[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    r += 2

    sheet[f"A{r}"] = "Note"
    sheet[f"A{r}"].font = note_font
    sheet[f"B{r}"] = ("The board is trimmed to a realistic depth per position -- "
                       "you won't see every player who could theoretically be "
                       "rostered, just anyone with a real shot at being drafted.")
    sheet[f"B{r}"].font = body_font
    sheet[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    r += 2

    sheet[f"A{r}"] = "Note"
    sheet[f"A{r}"].font = note_font
    sheet[f"B{r}"] = ("If you change anything on Setup mid-draft, every value "
                       "recalculates instantly -- but any Price Paid you've "
                       "already entered stays exactly as typed.")
    sheet[f"B{r}"].font = body_font
    sheet[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    r += 3

    sheet[f"A{r}"] = "Field Legend"
    sheet[f"A{r}"].font = legend_header_font
    r += 1

    legend = [
        ("Player / Team", "Who and where."),
        ("Proj Pts", "Projected season points, based on your selected Scoring Format."),
        ("Starting Value", "What we'd expect this player to cost before any picks are made."),
        ("Price Paid", "Enter the real winning bid once a player is drafted -- blank until then."),
        ("Market Value", "The live price estimate. Equals Price Paid once you've entered it; "
                          "otherwise it's the current estimate given who's already gone."),
    ]
    for label, desc in legend:
        sheet[f"A{r}"] = label
        sheet[f"A{r}"].font = body_font
        sheet[f"C{r}"] = desc
        sheet[f"C{r}"].font = body_font
        sheet[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    sheet.column_dimensions["A"].width = 21
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 55

    return sheet


def add_setup_sheet(wb):
    sheet = wb.create_sheet("Setup", 1)
    sheet["A1"] = "League Setup"
    sheet["A1"].font = Font(bold=True, size=13)

    thin = Side(style="thin", color="B0B0B0")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    input_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    def whole_dv(cell, lo, hi, label):
        dv = DataValidation(type="whole", operator="between", formula1=str(lo), formula2=str(hi),
                             allow_blank=False, showErrorMessage=True,
                             errorTitle=f"Invalid {label}", error=f"Enter a whole number ({lo}-{hi}).")
        sheet.add_data_validation(dv)
        dv.add(cell)

    def list_dv(cell, options, label):
        dv = DataValidation(type="list", formula1='"' + ",".join(str(o) for o in options) + '"',
                             allow_blank=False, showErrorMessage=True,
                             errorTitle=f"Invalid {label}", error=f"Choose one of: {', '.join(str(o) for o in options)}.")
        sheet.add_data_validation(dv)
        dv.add(cell)

    def input_row(row_num, label, value_cell, value, options=None, lo=None, hi=None):
        """One bordered, shaded [Label | Value] row -- reads as a table row."""
        sheet[f"A{row_num}"] = label
        sheet[value_cell] = value
        if options is not None:
            list_dv(value_cell, options, label)
        else:
            whole_dv(value_cell, lo, hi, label)
        sheet[f"A{row_num}"].border = box_border
        sheet[value_cell].border = box_border
        sheet[value_cell].fill = input_fill
        sheet[value_cell].alignment = Alignment(horizontal="center")

    input_row(3, "Starting Budget ($)", "B3", BUDGET, lo=15, hi=1000)
    input_row(5, "Number of Teams", "B5", TEAMS, options=TEAM_COUNTS)
    input_row(7, "Scoring Format", "B7", FORMAT_LABEL[DEFAULT_FORMAT],
              options=[FORMAT_LABEL[f] for f in FORMATS])

    sheet["A9"] = "Roster Shape"
    sheet["A9"].font = Font(bold=True, size=12)

    roster_rows = [
        ("QB Starters", "B11", 1, list(range(0, 3))),
        ("RB Starters", "B13", BASELINE_STARTERS["RB"], list(range(0, 4))),
        ("WR Starters", "B15", BASELINE_STARTERS["WR"], list(range(0, 5))),
        ("TE Starters", "B17", BASELINE_STARTERS["TE"], list(range(0, 3))),
    ]
    row_num = 11
    for label, cell, default, options in roster_rows:
        input_row(row_num, label, cell, default, options=options)
        row_num += 2

    # One independent count per flex TYPE (not a type-selector + count) --
    # matches the user's own cheat-sheet convention and lets multiple flex
    # types combine freely in the same league (e.g. 1 RB/WR + 1 WR/TE +
    # 1 Superflex all at once), which a single type-selector couldn't
    # represent.
    flex_rows = [
        ("Flex: RB/WR/TE Count", FLEX_SLOTS, "RB_WR_TE"),
        ("Flex: RB/WR Count", 0, "RB_WR"),
        ("Flex: WR/TE Count", 0, "WR_TE"),
        ("Superflex Count (QB/RB/WR/TE)", 0, "SUPERFLEX"),
    ]
    for label, default, key in flex_rows:
        input_row(row_num, label, f"B{row_num}", default, options=list(range(0, 7)))
        row_num += 2

    input_row(row_num, "DEF", f"B{row_num}", 1, options=[0, 1])
    row_num += 2

    input_row(row_num, "K", f"B{row_num}", 1, options=[0, 1])
    row_num += 2

    input_row(row_num, "Bench", f"B{row_num}", BENCH_SPOTS, options=list(range(3, 11)))
    row_num += 2

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 10

    return sheet, setup_cell_refs()


def setup_cell_refs():
    """The Setup sheet's input-cell addresses, as a standalone constant --
    entirely deterministic given add_setup_sheet()'s fixed row layout, so
    this can be used WITHOUT building/touching a Setup worksheet at all.
    Needed when preserving an existing, user-edited Setup tab on rebuild
    (see main()) -- the formulas built elsewhere still need to know which
    cells to reference even though this run isn't writing that sheet."""
    return {
        "qb_starters": "Setup!$B$11", "rb_starters": "Setup!$B$13",
        "wr_starters": "Setup!$B$15", "te_starters": "Setup!$B$17",
        "flex_counts": {
            "RB_WR_TE": "Setup!$B$19", "RB_WR": "Setup!$B$21",
            "WR_TE": "Setup!$B$23", "SUPERFLEX": "Setup!$B$25",
        },
        "def_count": "Setup!$B$27", "k_count": "Setup!$B$29",
        "bench_count": "Setup!$B$31",
    }


def build_draft_board_shell(wb, by_fmt, board_keys, layout):
    from name_match import normalize_name

    board = wb.create_sheet("Draft Board", 2)
    board.freeze_panes = "A2"

    points_by_fmt_pos = {}
    for fmt in FORMATS:
        for pos in POINTS_COL:
            d = {}
            for p in by_fmt[fmt][pos]:
                d[normalize_name(p["name"])] = p
            points_by_fmt_pos[(fmt, pos)] = d

    for pos in POINTS_COL:
        start = layout[pos]["start_col"]
        for j, header in enumerate(BLOCK_COLS):
            cell = board.cell(1, start + j, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="bottom")
        board.row_dimensions[1].height = 30
        fill = PatternFill(start_color=POSITION_FILL[pos], end_color=POSITION_FILL[pos], fill_type="solid")

        for i, key in enumerate(board_keys[pos], start=2):
            ref_player = next(points_by_fmt_pos[(fmt, pos)][key] for fmt in FORMATS
                               if key in points_by_fmt_pos[(fmt, pos)])
            board.cell(i, start + OFF_PLAYER, ref_player["name"]).fill = fill
            board.cell(i, start + OFF_TEAM, ref_player["team"]).fill = fill
            for fmt, off in [("std", OFF_PTS_STD), ("half_ppr", OFF_PTS_HALF), ("ppr", OFF_PTS_PPR)]:
                p = points_by_fmt_pos[(fmt, pos)].get(key)
                pts = round(p["blended_points"], 3) if p else 0.0
                board.cell(i, start + off, pts).fill = fill
            board.cell(i, start + OFF_ROWNUM, i).fill = fill  # stable tie-break id, literal not formula
            for off in (OFF_SEL_PTS, OFF_TIER, OFF_WVORP, OFF_GAP_START, OFF_GAPTIER_TOP,
                        OFF_TIER_START, OFF_STATIC, OFF_PRICE, OFF_LIVE):
                board.cell(i, start + off).fill = fill  # formulas/blank, filled in later

        # openpyxl's width unit isn't 1:1 with Excel's displayed character
        # width for the default font -- measured empirically (set X, Excel
        # showed X-0.71..0.72 on open), matches the standard 5/7 padding
        # constant for Calibri 11. WIDTH_FIX corrects for it so the values
        # below are what actually shows up in Excel, not what openpyxl
        # nominally stores.
        WIDTH_FIX = 5 / 7
        widths = {OFF_PLAYER: 18, OFF_TEAM: 7.57, OFF_SEL_PTS: 6.29, OFF_TIER: 7, OFF_WVORP: 13,
                  OFF_STATIC: 9.43, OFF_PRICE: 8.29, OFF_LIVE: 9.57}
        for offset, width in widths.items():
            board.column_dimensions[get_column_letter(start + offset)].width = width + WIDTH_FIX
        for off in (OFF_PTS_STD, OFF_PTS_HALF, OFF_PTS_PPR, OFF_TIER, OFF_WVORP,
                    OFF_ROWNUM, OFF_GAP_START, OFF_GAPTIER_TOP, OFF_TIER_START):
            board.column_dimensions[get_column_letter(start + off)].hidden = True
        board.column_dimensions[get_column_letter(start + BLOCK_WIDTH)].width = 2.0 + WIDTH_FIX

        last_row = layout[pos]["last_row"]
        price_col = get_column_letter(start + OFF_PRICE)
        dv = DataValidation(type="whole", operator="between", formula1="0", formula2="400",
                             allow_blank=True, showErrorMessage=True,
                             errorTitle="Invalid price", error="Enter a whole-dollar amount (0-400).")
        board.add_data_validation(dv)
        dv.add(f"{price_col}2:{price_col}{last_row}")

        # Selected Points: 3-way CHOOSE by Setup!Format
        fmt_list = "{" + ",".join(f'"{FORMAT_LABEL[f]}"' for f in FORMATS) + "}"
        std_col = get_column_letter(start + OFF_PTS_STD)
        half_col = get_column_letter(start + OFF_PTS_HALF)
        ppr_col = get_column_letter(start + OFF_PTS_PPR)
        for i in range(2, last_row + 1):
            board.cell(i, start + OFF_SEL_PTS).value = (
                f"=CHOOSE(MATCH(Setup!$B$7,{fmt_list},0),{std_col}{i},{half_col}{i},{ppr_col}{i})"
            )

        # Each position gets its own Table so the header row's built-in
        # filter/sort buttons only ever touch that position's rows -- a
        # plain worksheet AutoFilter is limited to one range per sheet,
        # which can't cover 4 independent position blocks. Sorting re-
        # orders each row's cells (formulas included) as a unit, so
        # relative same-row references (e.g. Tier depending on this row's
        # own Weighted VORP) stay correct; every cross-row formula in this
        # workbook uses absolute whole-column ranges, which don't care
        # about row order at all. Verified after build, not assumed.
        end_col = get_column_letter(start + BLOCK_WIDTH - 1)
        table_ref = f"{get_column_letter(start)}1:{end_col}{last_row}"
        table = Table(displayName=f"{pos}Board", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=False,
            showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
        )
        board.add_table(table)

    return board


def build_live_calc(wb, layout, setup_cells):
    calc = wb.create_sheet("Live Calc")
    row = 1

    # ── $ SETUP ────────────────────────────────────────────────────────
    calc.cell(row, 1, "$ SETUP").font = Font(bold=True)
    row += 1
    calc.cell(row, 1, "Total Teams")
    calc.cell(row, 2).value = "=Setup!$B$5"
    teams_row = row
    row += 1
    calc.cell(row, 1, "Total Roster Spots")
    all_flex_counts = "+".join(setup_cells["flex_counts"].values())
    roster_spots_formula = (
        f"=B{teams_row}*({setup_cells['qb_starters']}+{setup_cells['rb_starters']}"
        f"+{setup_cells['wr_starters']}+{setup_cells['te_starters']}"
        f"+({all_flex_counts})"
        f"+{setup_cells['def_count']}+{setup_cells['k_count']}+{setup_cells['bench_count']})"
    )
    calc.cell(row, 2).value = roster_spots_formula
    spots_row = row
    row += 1
    calc.cell(row, 1, "Total Money ($)")
    calc.cell(row, 2).value = f"=B{teams_row}*Setup!$B$3"
    money_row = row
    row += 1
    calc.cell(row, 1, "Total Discretionary ($)")
    calc.cell(row, 2).value = f"=B{money_row}-B{spots_row}*1"
    discretionary_row = row
    row += 2

    # ── ROSTER REGIME: has_superflex_regime = QB starters 2 OR superflex >=1
    calc.cell(row, 1, "ROSTER REGIME").font = Font(bold=True)
    row += 1
    calc.cell(row, 1, "Has Superflex-Level QB Demand")
    calc.cell(row, 2).value = f"=OR({setup_cells['qb_starters']}=2,{setup_cells['flex_counts']['SUPERFLEX']}>=1)"
    regime_row = row
    row += 2

    # ── BASELINE (real Teams/Format calibration) ──────────────────────
    calc.cell(row, 1, "BASELINE (real Teams/Format calibration)").font = Font(bold=True)
    row += 1
    for j, h in enumerate(["Position", "Baseline Rank", "Baseline Exponent",
                            "Baseline Budget Share", "SF-Shift Exp Ratio", "SF-Shift Share Ratio",
                            "Exponent Slope/Flex", "Budget-Share Slope/Flex"]):
        calc.cell(row, 1 + j, h).font = Font(bold=True)
    baseline_row = {}
    for pos in POINTS_COL:
        row += 1
        calc.cell(row, 1, pos)
        rank_cells = {ck: None for ck in CONFIG_KEYS}
        exp_cells = {ck: None for ck in CONFIG_KEYS}
        share_cells = {ck: None for ck in CONFIG_KEYS}
        for idx, ck in enumerate(CONFIG_KEYS):
            fmt, teams = ck
            cal = CALIBRATED[(teams, fmt)]
            col = 22 + idx * 3  # far right side area, 15*3=45 cols wide
            calc.cell(row, col, cal["ranks"][pos])
            calc.cell(row, col + 1, round(cal["exponents"][pos], 4))
            calc.cell(row, col + 2, round(cal["budget_share"][pos], 5))
            rank_cells[ck] = f"{get_column_letter(col)}{row}"
            exp_cells[ck] = f"{get_column_letter(col + 1)}{row}"
            share_cells[ck] = f"{get_column_letter(col + 2)}{row}"
        calc.cell(row, 2).value = f"={teamfmt_choose(rank_cells)}"
        calc.cell(row, 3).value = f"={teamfmt_choose(exp_cells)}"
        calc.cell(row, 4).value = f"={teamfmt_choose(share_cells)}"

        sf_anchor_baseline = CALIBRATED[(SUPERFLEX_ANCHOR["teams"], SUPERFLEX_ANCHOR["fmt"])]
        sf_exp_ratio = SUPERFLEX_ANCHOR["exponents"][pos] / sf_anchor_baseline["exponents"][pos]
        sf_share_ratio = SUPERFLEX_ANCHOR["budget_share"][pos] / sf_anchor_baseline["budget_share"][pos]
        calc.cell(row, 5, round(sf_exp_ratio, 4))
        calc.cell(row, 6, round(sf_share_ratio, 4))
        calc.cell(row, 7, round(EXPONENT_SLOPE_PER_FLEX[pos], 4))
        calc.cell(row, 8, round(BUDGET_SHARE_SLOPE_PER_FLEX[pos], 5))
        baseline_row[pos] = row
    row += 2

    # ── ROSTER-ADJUSTED (live formula) ────────────────────────────────
    calc.cell(row, 1, "ROSTER-ADJUSTED").font = Font(bold=True)
    row += 1
    for j, h in enumerate(["Position", "Demand (current)", "Demand (baseline)",
                            "Adjusted Rank", "Exponent (final)",
                            "Raw Adjusted Share", "Final Budget Share", "Replacement Points",
                            "Starter Demand Ratio"]):
        calc.cell(row, 1 + j, h).font = Font(bold=True)
    adj_row = {}
    # RB_WR_TE-type flex count only -- the exponent/budget-share slope is
    # only valid for that flex type (see roster_formula.py docstring); the
    # RB/WR-only and WR/TE-only counts don't feed it (no data for either).
    calc.cell(row - 1, 10, "RB/WR/TE-type Flex Count").font = Font(bold=True)
    calc.cell(row, 10).value = f"={setup_cells['flex_counts']['RB_WR_TE']}"
    rb_wr_te_flex_count_cell = f"$J${row}"

    for pos in POINTS_COL:
        row += 1
        calc.cell(row, 1, pos)

        starters_cell = setup_cells[{"QB": "qb_starters", "RB": "rb_starters",
                                       "WR": "wr_starters", "TE": "te_starters"}[pos]]

        # Demand = starters + SUM over all 4 independent flex-type counts
        # (each multiplied by that type's own share for this position --
        # constants known at build time, no CHOOSE/MATCH needed since each
        # flex type now has its own dedicated Setup cell, not a shared
        # type-selector). Multiple types combine freely by construction.
        flex_terms = "+".join(
            f"{round(FLEX_TYPES[ft].get(pos, 0.0), 4)}*{setup_cells['flex_counts'][ft]}"
            for ft in ("RB_WR_TE", "RB_WR", "WR_TE", "SUPERFLEX")
        )
        calc.cell(row, 2).value = f"={starters_cell}+{flex_terms}"

        baseline_starters = BASELINE_STARTERS[pos]
        baseline_flex_share = round(FLEX_TYPES["RB_WR_TE"].get(pos, 0.0), 4)
        calc.cell(row, 3).value = f"={baseline_starters}+{baseline_flex_share}*{FLEX_SLOTS}"

        # Clamped to >=1: a position can now have zero demand (0 starters,
        # no flex-eligibility) if the user configures it that way -- LARGE()
        # errors on k=0, so this floors it at "1 player of relevance"
        # rather than erroring. Not a meaningful real-world case, just a
        # safety net against a genuinely empty position.
        # Rank = baseline_rank * flex_ratio * starter_ratio^DAMPING.
        # Mirrors roster_formula.py: the flex dimension stays fully
        # proportional (the real 3-flex/superflex anchors require that to
        # round-trip), while the starter dimension is damped -- fitted
        # against a real FantasyPros 12-team 3WR-vs-2WR A/B. `denom` is
        # "baseline starters carrying THIS shape's flex", the shared
        # midpoint that splits the two dimensions cleanly.
        denom = f"({baseline_starters}+B{row}-{starters_cell})"
        flex_ratio = f"IFERROR({denom}/C{row},1)"
        calc.cell(row, 4).value = (
            f"=MAX(1,ROUND(B{baseline_row[pos]}*{flex_ratio}"
            f"*I{row}^{STARTER_RANK_DAMPING},0))"
        )

        regime_cell = f"$B${regime_row}"
        flex_delta = f"({rb_wr_te_flex_count_cell}-{FLEX_SLOTS})"
        exp_interp = f"C{baseline_row[pos]}+G{baseline_row[pos]}*{flex_delta}"
        share_interp = f"D{baseline_row[pos]}+H{baseline_row[pos]}*{flex_delta}"

        # Starter-demand ratio -- mirrors roster_formula.py's
        # _scale_share_by_starter_demand(). Without this, changing a
        # starter count moved the replacement line (col D) but NOT the
        # money, so the pool collapsed while the position kept its full
        # budget share and the rate exploded (real user-reported bug:
        # WR 3->2 starters sent Puka Nacua UP from $52 to $83). See that
        # function's docstring for the full mechanism.
        #
        # Denominator rebuilds "baseline starters + THIS shape's flex" as
        # B - current_starters + baseline_starters, deliberately holding
        # flex constant on both sides so this isolates the starters
        # dimension only -- the flex dimension is already covered by the
        # H-column slope in share_interp, and a full demand ratio would
        # double-count it.
        calc.cell(row, 9).value = (
            f"=IFERROR(B{row}/({baseline_starters}+B{row}-{starters_cell}),1)"
        )
        calc.cell(row, 5).value = f"=IF({regime_cell},C{baseline_row[pos]}*E{baseline_row[pos]},{exp_interp})"
        # Only the non-regime branch gets the starter scaling, matching
        # roster_formula.py: the superflex regime uses REAL measured
        # shift ratios, which already price in that shape's own demand.
        calc.cell(row, 6).value = (
            f"=IF({regime_cell},D{baseline_row[pos]}*F{baseline_row[pos]},({share_interp})*I{row})"
        )

        pts_rng = block_range(layout, pos, OFF_SEL_PTS)
        calc.cell(row, 8).value = f"=LARGE({pts_rng},MIN(D{row},{layout[pos]['n_rows']}))"

        adj_row[pos] = row
    # Final Budget Share needs renormalization across positions -> second pass
    for pos in POINTS_COL:
        r = adj_row[pos]
        total_raw_share = "+".join(f"F{adj_row[p]}" for p in POINTS_COL)
        calc.cell(r, 7).value = f"=F{r}/({total_raw_share})"
    row += 2

    # ── POSITION RATES ─────────────────────────────────────────────────
    calc.cell(row, 1, "POSITION RATES").font = Font(bold=True)
    row += 1
    for j, h in enumerate(["Position", "Discretionary $", "Total Weighted VORP", "Rate"]):
        calc.cell(row, 1 + j, h).font = Font(bold=True)
    rate_row = {}
    for pos in POINTS_COL:
        row += 1
        calc.cell(row, 1, pos)
        calc.cell(row, 2).value = f"=$B${discretionary_row}*G{adj_row[pos]}"
        wvorp_rng = block_range(layout, pos, OFF_WVORP)
        calc.cell(row, 3).value = f"=SUM({wvorp_rng})"
        # Guard against a near-empty (not exactly zero) Total Weighted VORP
        # pool: if a position is configured with very little real demand
        # (e.g. 0 starters, only lightly flex-eligible), Adjusted Rank can
        # land very shallow, leaving only a handful of players with any
        # positive weighted VORP -- and dividing a real discretionary $
        # amount by that tiny pool produces an absurdly large rate.
        # Confirmed with real numbers, not guessed: a genuine 0-TE-starter/
        # 1-flex config produced TotalWeightedVORP=1.89 (a NORMAL
        # position's pool runs in the hundreds to thousands), Rate=104,
        # and a single player at $198. A threshold of 1.0 didn't catch
        # this (1.89 > 1) -- widened to 20, comfortably below every real
        # baseline pool size in this project but well above what a
        # genuinely-degenerate pool produces. Treating anything under the
        # threshold the same as exactly zero (Rate=0, everyone in that
        # position floors to the $1 static value) is the sensible outcome
        # for a position with essentially no real demand.
        calc.cell(row, 4).value = f"=IF(C{row}<20,0,IFERROR(B{row}/C{row},0))"
        rate_row[pos] = row
    row += 2

    # ── Per-position TIER tables (gap-detection tiers, see module
    # docstring "Round 8 fourth extension: gap-detection tiers"). Tier
    # count is data-dependent (unlike the old fixed-bucket scheme), so the
    # table is sized to n_rows -- the true worst case (every player starts
    # a new tier) -- guaranteeing it can never be too small regardless of
    # Setup, which is what actually matters: this project has hit the
    # "#N/A because the lookup table was too small" failure twice already
    # in this same round, so this sizing is deliberately generous rather
    # than clever.
    tier_table_range = {}
    for pos in POINTS_COL:
        calc.cell(row, 1, f"{pos} tiers").font = Font(bold=True)
        row += 1
        for j, h in enumerate(["Tier", "Drafted Price Sum", "Drafted Static Sum",
                                "Drafted Count", "Tier Avg Static", "Dampening", "Tier Factor"]):
            calc.cell(row, 1 + j, h).font = Font(bold=True)
        table_start = row + 1
        price_rng = block_range(layout, pos, OFF_PRICE)
        static_rng = block_range(layout, pos, OFF_STATIC)
        tier_rng = block_range(layout, pos, OFF_TIER)
        n_rows = layout[pos]["n_rows"]
        for t in range(1, n_rows + 1):
            row += 1
            calc.cell(row, 1, t)
            calc.cell(row, 2).value = f'=SUMIFS({price_rng},{tier_rng},A{row},{price_rng},"<>")'
            calc.cell(row, 3).value = f'=SUMIFS({static_rng},{tier_rng},A{row},{price_rng},"<>")'
            calc.cell(row, 4).value = f'=COUNTIFS({tier_rng},A{row},{price_rng},"<>")'
            # Tier Avg Static: every player in the tier (drafted or not),
            # not just drafted ones -- this is what the Dampening column
            # measures against, so a tier full of $1-floor players stays
            # dampened even before anyone in it is drafted.
            calc.cell(row, 5).value = f'=IFERROR(AVERAGEIFS({static_rng},{tier_rng},A{row}),0)'
            # Dampening: tiers with real dollar value get the full
            # re-rating signal; tiers near the $1 floor barely move even
            # once the sample gate opens. See CLAUDE.md "Round 8".
            calc.cell(row, 6).value = (
                f"=MIN(1,E{row}/(Setup!$B$3*{TIER_DAMPEN_BUDGET_FRACTION}))"
            )
            # A single drafted player in a tier is one bidder's opinion on
            # one player, not a market signal -- require at least
            # TIER_MIN_SAMPLE drafted players before letting the tier
            # re-rate its (still-undrafted) neighbors at all. Below that,
            # factor stays at 1.0 regardless of how anomalous that one
            # price was. See CLAUDE.md "Round 8".
            calc.cell(row, 7).value = (
                f"=IF(D{row}<{TIER_MIN_SAMPLE},1,"
                f"MEDIAN({TIER_FACTOR_MIN},{TIER_FACTOR_MAX},"
                f"1+(IFERROR(B{row}/C{row},1)-1)*F{row}))"
            )
        table_end = row
        tier_table_range[pos] = (f"$A${table_start}:$A${table_end}", f"$G${table_start}:$G${table_end}")
        row += 2

    # ── GLOBAL (whole-draft) depletion table ──────────────────────────
    calc.cell(row, 1, "GLOBAL (whole draft)").font = Font(bold=True)
    row += 1
    g_headers = ["Initial Discretionary $", "Initial Weighted VORP", "Initial Rate",
                 "Drafted $ (all positions)", "Drafted Count (all positions)",
                 "Drafted Weighted VORP (all positions)",
                 "Remaining Discretionary $", "Remaining Weighted VORP",
                 "Current Rate", "Depletion Factor"]
    for j, h in enumerate(g_headers):
        calc.cell(row, 1 + j, h).font = Font(bold=True)
    g_row = row + 1

    calc.cell(g_row, 1).value = f"=B{discretionary_row}"
    rate_cells = "+".join(f"C{rate_row[pos]}" for pos in POINTS_COL)
    calc.cell(g_row, 2).value = f"={rate_cells}"
    calc.cell(g_row, 3).value = f"=IFERROR(A{g_row}/B{g_row},0)"

    price_terms = "+".join(f"SUM({block_range(layout, pos, OFF_PRICE)})" for pos in POINTS_COL)
    count_terms = "+".join(f"COUNT({block_range(layout, pos, OFF_PRICE)})" for pos in POINTS_COL)
    wvorp_terms = "+".join(
        f'SUMIFS({block_range(layout, pos, OFF_WVORP)},{block_range(layout, pos, OFF_PRICE)},"<>")'
        for pos in POINTS_COL
    )
    calc.cell(g_row, 4).value = f"={price_terms}"
    calc.cell(g_row, 5).value = f"={count_terms}"
    calc.cell(g_row, 6).value = f"={wvorp_terms}"
    calc.cell(g_row, 7).value = f"=A{g_row}-(D{g_row}-E{g_row})"
    calc.cell(g_row, 8).value = f"=B{g_row}-F{g_row}"
    calc.cell(g_row, 9).value = f"=IFERROR(G{g_row}/H{g_row},0)"
    calc.cell(g_row, 10).value = (
        f"=MEDIAN({GLOBAL_FACTOR_MIN},{GLOBAL_FACTOR_MAX},IFERROR(I{g_row}/C{g_row},1))"
    )

    for col, width in zip("ABCDEFGHIJ", (16, 20, 20, 16, 22, 18, 18, 16, 20, 18)):
        calc.column_dimensions[col].width = width

    return calc, adj_row, rate_row, tier_table_range, g_row


def main():
    by_fmt = compute_all_points()
    board_keys = board_player_list(by_fmt)
    layout = build_block_layout(board_keys)

    # Instructions and Setup are user-owned once they exist -- the user
    # edits them by hand (moved/reworded content, column widths, etc.) and
    # explicitly does not want this script overwriting either. So: only
    # build them fresh the very first time this file doesn't exist yet;
    # every rebuild after that preserves them untouched and only replaces
    # Draft Board / Live Calc / My Team.
    if OUT_FILE.exists():
        wb = load_workbook(OUT_FILE)
        for name in ("Draft Board", "Live Calc", "My Team"):
            if name in wb.sheetnames:
                del wb[name]
        setup_cells = setup_cell_refs()
    else:
        wb = Workbook()
        wb.remove(wb.active)
        add_instructions_sheet(wb)
        _, setup_cells = add_setup_sheet(wb)

    board = build_draft_board_shell(wb, by_fmt, board_keys, layout)
    calc, adj_row, rate_row, tier_table_range, g_row = build_live_calc(wb, layout, setup_cells)

    global_factor_cell = f"'Live Calc'!$J${g_row}"

    for pos in POINTS_COL:
        start = layout[pos]["start_col"]
        tier_range, factor_range = tier_table_range[pos]
        sel_pts_col = get_column_letter(start + OFF_SEL_PTS)
        tier_col = get_column_letter(start + OFF_TIER)
        wvorp_col = get_column_letter(start + OFF_WVORP)
        rownum_col = get_column_letter(start + OFF_ROWNUM)
        gap_start_col = get_column_letter(start + OFF_GAP_START)
        gaptier_top_col = get_column_letter(start + OFF_GAPTIER_TOP)
        tier_start_col = get_column_letter(start + OFF_TIER_START)
        static_col = get_column_letter(start + OFF_STATIC)
        price_col = get_column_letter(start + OFF_PRICE)
        live_col = get_column_letter(start + OFF_LIVE)
        rate_cell = f"'Live Calc'!$D${rate_row[pos]}"
        replacement_cell = f"'Live Calc'!$H${adj_row[pos]}"
        exponent_cell = f"'Live Calc'!$E${adj_row[pos]}"
        wvorp_rng = block_range(layout, pos, OFF_WVORP)
        rownum_rng = block_range(layout, pos, OFF_ROWNUM)
        gap_start_rng = block_range(layout, pos, OFF_GAP_START)
        tier_start_rng = block_range(layout, pos, OFF_TIER_START)
        n_rows = layout[pos]["n_rows"]
        for i in range(2, layout[pos]["last_row"] + 1):
            board[f"{wvorp_col}{i}"].value = (
                f"=MAX(0,{sel_pts_col}{i}-{replacement_cell})^{exponent_cell}"
            )
            # Gap Start: pure neighbor gap-detection (see module docstring
            # "Round 8 fourth extension: gap-detection tiers") -- a real
            # replacement for a value-proportional-band version that
            # shipped and was wrong (user: "why is there no tier 1 in any
            # position? these tiers went too far the other way" -- close
            # players like Gibbs/Bijan were being split by cumulative-share
            # bucket boundaries that had nothing to do with any actual gap
            # between them).
            #
            # A player "starts a new tier" if MINIFS(">"&own) -- the
            # Weighted VORP of whoever is ranked immediately above them --
            # drops by more than GAP_THRESHOLD relatively, or if there's
            # no one above at all (the single most valuable player).
            # MINIFS(">"&own) returns 0 when no cell qualifies, which
            # cleanly identifies "no one above" since real Weighted VORP
            # is never negative.
            #
            # Ties need special handling: many players legitimately share
            # the exact same Weighted VORP (most commonly 0, at or below
            # replacement -- confirmed most positions have 100+ tied
            # zero-value players). MINIFS(">"&own) skips over ties to the
            # next DISTINCT value, so every tied player would otherwise
            # see the same "gap" to that next distinct value and each
            # independently flag as a tier start -- fragmenting what
            # should be one shared tier into dozens. Fixed by only
            # evaluating the gap check for the CANONICAL representative of
            # each tied group (the one with the smallest Row Num, a stable
            # build-time id, not a live position) -- every other member of
            # a tie gets Gap Start = 0 and inherits the same tier via the
            # COUNTIFS pattern below, since it evaluates identically for
            # every player sharing that same own_wvorp value.
            # _xlfn. prefix required: MINIFS is Excel 2019+, same class of
            # bug as the earlier RANK.EQ issue -- openpyxl writes formula
            # text directly into the XLSX XML without adding the prefix
            # Excel's own UI adds silently, so this shows #NAME? on open
            # otherwise (confirmed via the same win32com error code check
            # as before, not guessed).
            board[f"{gap_start_col}{i}"].value = (
                f'=IF({rownum_col}{i}<>_xlfn.MINIFS({rownum_rng},{wvorp_rng},"="&{wvorp_col}{i}),0,'
                f'IF(_xlfn.MINIFS({wvorp_rng},{wvorp_rng},">"&{wvorp_col}{i})=0,1,'
                f'IF((_xlfn.MINIFS({wvorp_rng},{wvorp_rng},">"&{wvorp_col}{i})-{wvorp_col}{i})'
                f'/_xlfn.MINIFS({wvorp_rng},{wvorp_rng},">"&{wvorp_col}{i})>{GAP_THRESHOLD},1,0)))'
            )
            # Gap Tier Top: the Weighted VORP of this player's own gap-tier
            # starter -- the smallest Gap-Start=1 value that's still >= my
            # own value (gap-tier starters partition the position into
            # non-overlapping ranges, so this always resolves to exactly
            # my own group's starter, never a neighboring group's).
            board[f"{gaptier_top_col}{i}"].value = (
                f'=_xlfn.MINIFS({wvorp_rng},{wvorp_rng},">="&{wvorp_col}{i},{gap_start_rng},1)'
            )
            # Tier Start (refined): see module docstring "Round 8 sixth
            # extension: cumulative tier-width cap". A long chain of small
            # individual gaps (each under GAP_THRESHOLD) can still add up
            # to a wide total range with no internal boundary -- this adds
            # a second trigger: split wherever my own "width bucket"
            # (how many MAX_TIER_WIDTH-sized steps I am below my gap-tier's
            # own top) differs from the width bucket of whoever's ranked
            # immediately above me. Compared to a bucket boundary rather
            # than a raw threshold check specifically so a long chain
            # splits into multiple correctly-sized groups instead of every
            # player past the first cap independently flagging as its own
            # 1-player tier. Same canonical-tie handling as Gap Start,
            # since two tied players get identical own/prev-neighbor
            # buckets and would otherwise both flag independently.
            own_bucket = (
                f'IFERROR(FLOOR(({gaptier_top_col}{i}-{wvorp_col}{i})/{gaptier_top_col}{i}'
                f"/{MAX_TIER_WIDTH},1),0)"
            )
            prev_bucket = (
                f'IFERROR(FLOOR(({gaptier_top_col}{i}-_xlfn.MINIFS({wvorp_rng},{wvorp_rng},">"&{wvorp_col}{i}))'
                f"/{gaptier_top_col}{i}/{MAX_TIER_WIDTH},1),0)"
            )
            board[f"{tier_start_col}{i}"].value = (
                f'=IF({rownum_col}{i}<>_xlfn.MINIFS({rownum_rng},{wvorp_rng},"="&{wvorp_col}{i}),0,'
                f"IF(OR({gap_start_col}{i}=1,{own_bucket}<>{prev_bucket}),1,0))"
            )
            # Tier: cumulative count of Tier Start flags among players
            # ranked at-or-above this one. This is a running COUNT, not a
            # bucket-index computation, so tier numbers are guaranteed
            # consecutive (1, 2, 3, ...) with no skips regardless of how
            # large any individual gap is -- fixes a second real bug the
            # user caught in the previous (bucket-based) design: "Why are
            # there missing tiers in QB, I would think it would also go in
            # incremental order no matter what the separation is."
            board[f"{tier_col}{i}"].value = (
                f'=COUNTIFS({wvorp_rng},">="&{wvorp_col}{i},{tier_start_rng},1)'
            )
            board[f"{static_col}{i}"].value = f"=ROUND(1+{wvorp_col}{i}*{rate_cell},0)"
            board[f"{live_col}{i}"].value = (
                f'=IF({price_col}{i}<>"",{price_col}{i},'
                f"ROUND({static_col}{i}"
                f"*INDEX('Live Calc'!{factor_range},MATCH({tier_col}{i},'Live Calc'!{tier_range},0))"
                f"*{global_factor_cell},0))"
            )

    add_my_team_sheet(wb, setup_cells)

    # Live Calc is the calculation engine (tier tables, global depletion
    # block) -- customers never need to see or touch it. Hidden, not
    # deleted, since every visible formula still references it.
    calc.sheet_state = "hidden"

    wb.active = wb.sheetnames.index("Instructions")

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")
    print(f"Change Setup!B3 (Budget), Setup!B5 (Teams), Setup!B7 (Format), or any Roster Shape "
          f"field to instantly recalculate every value.")
    print("Enter real prices in each position's 'Price Paid ($)' column on Draft Board.")
    print("Board is trimmed to top " +
          ", ".join(f"{BOARD_DEPTH[p]} {p}" for p in POINTS_COL) +
          " (union of all 3 formats' own top-N by points).")
    print("Track your own roster/budget on the My Team sheet (independent of Draft Board).")


def add_my_team_sheet(wb, setup_cells):
    """Simple personal budget tracker: enter who you drafted and what you
    paid, get real-time remaining budget / max next bid / average $ left
    per remaining spot. Completely independent of Draft Board/Live Calc's
    recalculation math. Roster slot list now reflects the LIVE Roster
    Shape inputs (QB/RB/WR/TE starters, flex, superflex, DEF, K, bench)
    instead of a fixed constant -- rebuilt as a formula-driven count per
    slot type rather than literal rows, since the exact starter counts are
    now variable."""
    # Generous upper bound across the whole range: QB2+RB3+WR4+TE2 starters
    # + 4 independent flex types x 6 each + DEF1+K1+Bench10, plus buffer.
    max_slots = (2 + 3 + 4 + 2 + 6 * 4 + 1 + 1 + 10) + 2
    last_row = max_slots + 1

    sheet = wb.create_sheet("My Team")
    sheet.append(["Roster Slot", "Player", "Price Paid ($)"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Slot labels are generated live from Setup's roster counts via a
    # simple repeating formula per row: this row's label is blank once
    # past the total configured roster size for the CURRENT settings.
    slot_defs = [
        ("QB", setup_cells["qb_starters"]), ("RB", setup_cells["rb_starters"]),
        ("WR", setup_cells["wr_starters"]), ("TE", setup_cells["te_starters"]),
        ("FLEX (RB/WR/TE)", setup_cells["flex_counts"]["RB_WR_TE"]),
        ("FLEX (RB/WR)", setup_cells["flex_counts"]["RB_WR"]),
        ("FLEX (WR/TE)", setup_cells["flex_counts"]["WR_TE"]),
        ("SUPERFLEX", setup_cells["flex_counts"]["SUPERFLEX"]),
        ("DEF", setup_cells["def_count"]), ("K", setup_cells["k_count"]),
        ("BENCH", setup_cells["bench_count"]),
    ]
    # Cumulative-count formula per row: label = the slot type whose
    # cumulative range this row index falls into, else "" (counts are
    # live/variable, so this can't be precomputed in Python -- built once
    # as a nested IF chain, evaluated per row).
    cum_terms = []
    running = "0"
    for label, cell in slot_defs:
        cum_terms.append((label, running, f"({running}+{cell})"))
        running = f"({running}+{cell})"
    nested = '""'
    for label, lo, hi in reversed(cum_terms):
        nested = f'IF(AND((ROW()-1)>{lo},(ROW()-1)<={hi}),"{label}",{nested})'
    for r in range(2, last_row + 1):
        formula = nested.replace("ROW()", str(r))
        sheet.cell(r, 1).value = f"={formula}"

    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="400",
                         allow_blank=True, showErrorMessage=True,
                         errorTitle="Invalid price", error="Enter a whole-dollar amount (0-400).")
    sheet.add_data_validation(dv)
    dv.add(f"C2:C{last_row}")

    # Summary block sits beside the roster list (columns E:F), not stacked
    # below it -- was originally placed at last_row+3 in columns A:B, which
    # put it many rows down for a full-depth roster and was easy to miss.
    # Fixed position matching the user's own manual layout: title in a
    # merged E2:F2 cell, an all-bordered [label|value] block in E3:F10.
    thin = Side(style="thin", color="000000")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells("E2:F2")
    sheet["E2"] = "My Team Tracker"
    sheet["E2"].font = Font(bold=True)
    sheet["E2"].border = box_border
    sheet["F2"].border = box_border

    total_spots_formula = "+".join(f"{cell}" for _, cell in slot_defs)
    rows = [
        ("Total Budget ($)", "=Setup!$B$3"),
        ("Roster Spots (total)", f"={total_spots_formula}"),
        ("Spots Filled", f'=COUNTIFS(A2:A{last_row},"<>",C2:C{last_row},"<>")'),
        ("Spots Remaining", None),
        ("Total Spent ($)", f"=SUM(C2:C{last_row})"),
        ("Money Remaining ($)", None),
        ("Max Bid on Next Player ($)", None),
        ("Avg $ per Remaining Spot ($)", None),
    ]
    r0 = 3
    for offset, (label, value) in enumerate(rows):
        r = r0 + offset
        sheet.cell(r, 5, label).font = Font(bold=True)
        sheet.cell(r, 5).border = box_border
        sheet.cell(r, 6).border = box_border
        if value is not None:
            sheet.cell(r, 6).value = value
    budget_r, spots_r, filled_r, remain_r, spent_r, money_r, max_r, avg_r = range(r0, r0 + 8)
    sheet.cell(remain_r, 6).value = f"=F{spots_r}-F{filled_r}"
    sheet.cell(money_r, 6).value = f"=F{budget_r}-F{spent_r}"
    sheet.cell(max_r, 6).value = f'=IF(F{remain_r}>0,F{money_r}-(F{remain_r}-1)*1,"Roster full")'
    sheet.cell(avg_r, 6).value = f'=IFERROR(F{money_r}/F{remain_r},"Roster full")'

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 26
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 3
    sheet.column_dimensions["E"].width = 26
    sheet.column_dimensions["F"].width = 15


if __name__ == "__main__":
    main()
